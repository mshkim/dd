import openpyxl
import xlrd
import re
from datetime import datetime
import collections
import xlwings
ITEMNUMBER = 100 # 아이템 개수 전역 변수로... 재입고필요한 상품을 위한

# 특정 상품을 구매시 무료로 제공하는 특정 상품 리스트 구한다. (나이키, 조던, 발렌시아 슈즈 악세사리 목록)
nikeShoesList = ['A005', 'A006', 'A007', 'A008', 'A009', 'A010', 'A011', 'A012', 'A013', 'A014', 'A015', 'A016']
adidasShoesList = ['A028', 'A029', 'A030', 'A031']
balenciagaShoesList = ['A047', 'A048', 'A049', 'A050']
freeBoxList = ['A019', 'A020', 'A018'] # 나이키, 아디다스, 발렌시아가 순

class OnlineShopTask():
    def __init__(self):
        self.orderList = 0
        self.packageNumber = 0
        self.naverProcessShipment = 0
        self.manageStock = 0

    # 네이버배송일괄 처리 엑셀
    def NaverPackageDelivery(self, orderListExel, packageNmnExel):
        self.orderList = openpyxl.load_workbook(orderListExel)
        self.packageNumber = xlrd.open_workbook(packageNmnExel)
        try:
            self.naverProcessShipment = openpyxl.load_workbook('네이버일괄발송양식.xlsx')
        except:
            print('네이버일괄발송양식 파일이 없습니다.')
        self.orderListList = []
        self.packageNumberList = []
        cntCell = 0

        getActiveSheepOL = self.orderList.active
        getActiveSheepPN = self.packageNumber.sheet_by_index(0)
        getActiveSheepNPS = self.naverProcessShipment.active

        # 네이버 배송 일괄 처리 엑셀에 상품주문번호, 택배방법, 택배사 입력
        # 네이버주문리스트 JOIN 택배송장 에 필요한 네이버주문리스트에서 정보 추출(orderListList)
        for row in getActiveSheepOL.rows:
            cntCell += 1
            if cntCell >= 3:
                orderNumber = row[0].value # 상품주문번호
                shippingMethod = row[2].value # 배송방법
                addressee = row[9].value # 수취인명
                contactAddress = row[25].value # 수취인연락처

                contactAddressR = re.sub('-', '', contactAddress) # 연락처 '-' 문자 제거

                newList = []
                newList.append(orderNumber)
                newList.append(addressee)
                newList.append(contactAddressR)
                self.orderListList.append(newList)

                getActiveSheepNPS.cell(row=cntCell - 1, column=1).value = orderNumber
                getActiveSheepNPS.cell(row=cntCell - 1, column=2).value = shippingMethod
                getActiveSheepNPS.cell(row=cntCell - 1, column=3).value = 'CJ대한통운'

        # 네이버주문리스트 JOIN 택배송장 에 필요한 택배송장에서 정보 추출(packageNumberList)
        num_rows = getActiveSheepPN.nrows  # 줄 수 가져오기
        for i in range(1, num_rows):
            billingNumber = getActiveSheepPN.cell_value(i, 7) # 운송장번호
            addressee = getActiveSheepPN.cell_value(i, 20) # 받는분
            contactAddress = getActiveSheepPN.cell_value(i, 21) # 받는분 전화번호

            newList = []
            newList.append(billingNumber)
            newList.append(addressee)
            newList.append(contactAddress)
            self.packageNumberList.append(newList)

        # 네이버주문리스트 JOIN 택배송장 --> 네이버일괄주문.송장번호 저장
        cntCell = 0
        for i in self.orderListList:
            for j in self.packageNumberList:
                if i[1] == j[1] and i[2] == j[2]:
                    cntCell += 1
                    getActiveSheepNPS.cell(row=cntCell+1, column=4).value = j[0]
                    #print(cntCell, i[0], i[1], i[2], j[1], j[2], j[0])
                    print(cntCell, '개 주문 확인 완료')

        date = datetime.today().strftime("%Y-%m-%d")
        fileName = orderListExel.replace(".xlsx", "") + ' 네이버일괄배송.xlsx'
        self.naverProcessShipment.save(filename=fileName)
        print('<네이버일괄배송 정보 엑셀 저장 완료>')

    # 재고 관리 엑셀
    def ManageItems(self, orderListExel, stockManageExel):
        self.orderList = openpyxl.load_workbook(orderListExel)
        self.manageStock = openpyxl.load_workbook(stockManageExel)
        naverOrderListCode = [] #네이버 쇼핑몰에 등록 된 상품리스트에 대한 코드
        stockCode = [] #전체 상품 코드
        cntCell = 0 # Control Cell RowIndex

        getActiveSheepOL = self.orderList.active
        getActiveSheepMS = self.manageStock['상품코드관리']

        # 네이버 주문 리스트 목록 상품명과 옵셥이 같은 상품코드관리 일치정보 추출
        for cellOL in getActiveSheepOL.rows:
            cntCell += 1
            if cntCell >= 3:
                naverStockName = cellOL[13].value
                naverStockOption = cellOL[15].value
                naverstockNum = int(cellOL[16].value)
                print(cntCell, '================================================')
                for cellMS in getActiveSheepMS:
                    manageStockName = cellMS[0].value
                    manageStockOption = cellMS[1].value
                    manageStockCode = cellMS[2].value
                    if naverStockName == manageStockName and naverStockOption == manageStockOption:
                        # 판매 수량 만큼 카운트
                        for i in range(1, naverstockNum+1):
                            naverOrderListCode.append(manageStockCode)
                        print(cntCell, cellMS[0].value, cellMS[1].value, cellMS[2].value)
                        #print(naverOrderListCode)

        # 무료 상품 처리
        freeItemCode1, num1 = self.ManageFreeItem(orderListExel, stockManageExel, nikeShoesList, freeBoxList[0])
        freeItemCode2, num2 = self.ManageFreeItem(orderListExel, stockManageExel, adidasShoesList, freeBoxList[1])
        freeItemCode3, num3 = self.ManageFreeItem(orderListExel, stockManageExel, balenciagaShoesList, freeBoxList[2])
        print(freeItemCode1, '상품 무료 지급 개수 :', num1)
        print(freeItemCode2, '상품 무료 지급 개수 :', num2)
        print(freeItemCode3, '상품 무료 지급 개수 :', num3)

        #전체 상품 코드 별 주문 개수 카운트
        cntCell2 = 0
        countOrderList = [] # 카운트된 리스트
        print('<상품코드별 개수 상황>')
        for cell in getActiveSheepMS:
            cntCell2 += 1
            code = cell[6].value
            if cntCell2 >= 2 and type(code) == str:
                if code == freeItemCode1:
                    countOrderList.append(naverOrderListCode.count(code) + num1)
                    print(code, naverOrderListCode.count(code) + num1)
                elif code == freeItemCode2:
                    countOrderList.append(naverOrderListCode.count(code) + num2)
                    print(code, naverOrderListCode.count(code) + num2)
                elif code == freeItemCode3:
                    countOrderList.append(naverOrderListCode.count(code) + num3)
                    print(code, naverOrderListCode.count(code) + num3)
                else:
                    countOrderList.append(naverOrderListCode.count(code))
                    print(code, naverOrderListCode.count(code))

        fileName = 0
        for i, j in enumerate(orderListExel.split('/')):
            if i == len(orderListExel.split('/')) - 1:
                fileName = j
                #print(i, j)

        # 파일이름에서 일자 추출
        fileName = fileName[2:10]
        fileName = fileName.replace("-", ".", 2)
        # 엑셀 시트 이름 출력 XX.XX. 형태
        sheetName = fileName[0:6]

        getActiveSheepMS = self.manageStock[sheetName]
        cntCell2 = 0
        dateColumn = 0
        for column in getActiveSheepMS[2]:
            cntCell2 += 1
            if fileName == column.value:
                dateColumn = cntCell2

        #당일 배송 시(당일 최소 배송 리스트) 해당 날짜에 바로 입력
        if getActiveSheepMS[3][dateColumn-1].value == None:
            for i, num in enumerate(countOrderList):
                getActiveSheepMS.cell(i+3, dateColumn).value = num

            # 상품 코드에 해당하는 상품명, 옵션명 찾아서 해당

        # 추가 배송 시(최소 배송 이후 추가적인 배송 발생경우) 당일 배송 + 추가 배송
        else:
            for i, num in enumerate(countOrderList):
                getActiveSheepMS.cell(i + 3, dateColumn).value = getActiveSheepMS[i+3][dateColumn - 1].value + num
        try:
            self.manageStock.save(filename=stockManageExel)
            print('<재고파일 업데이트 완료>')
        except:
            print('재고관리 파일 닫고 다시 실행.')

        return 0

    # 특정 상품 특정 개수 이상 구매시 무료 상품 지급관련 재고 카운팅 & 재고관리 파일에 업데이트
    def ManageFreeItem(self, orderListExel, stockManageExel, shoesList, freeBox):
        self.orderList = openpyxl.load_workbook(orderListExel)
        self.manageStock = openpyxl.load_workbook(stockManageExel)
        getActiveSheepOL = self.orderList.active
        getActiveSheepMS = self.manageStock['상품코드관리']

        MS = []
        # 재고관리.상품코드관리 ;; [상품명, 옵션명, 상품코드]
        for code in shoesList:
            list = [] #[상품명, 옵션명, 상품코드] create tuple
            for row in getActiveSheepMS.rows:
                if code == row[6].value:
                    list.append(row[4].value)
                    list.append(row[5].value)
                    list.append(row[6].value)
            MS.append(list)

        sellShoes = []
        # 네이버주문리스트 ;; [주문자, 핸드폰 번호, 상품명, 옵션정보, 구매개수] 조인
        for row in getActiveSheepOL:
            list = []
            for i in MS:
                if i[0] == row[13].value and i[1] == row[15].value:
                    if row[16].value == 1:
                        list.append(row[7].value + row[25].value) # 이름
                        #list.append(row[25].value) # 번호
                        #list.append(row[16].value)# 개수
                        sellShoes.extend(list)
                    else:
                        for i in range(0, int(row[16].value)):
                            list = []
                            list.append(row[7].value + row[25].value)  # 이름
                            #list.append(row[25].value)  # 번호
                            #list.append(1)  # 개수
                            sellShoes.extend(list)

        print('<', freeBox, '상품 무료 지급 대상자 이름, 번호 목록>')
        for i in sellShoes:
            print(i)

        print('==================================================')

        counting = collections.Counter(sellShoes)
        freeBoxCount = 0
        for i in counting.values():
            if i >= 2:
                freeBoxCount += int(i / 2)

        #print(freeBox, freeBoxCount)

        # 무료 상품 코드 & 개수 반환 ex) A019 4
        return freeBox, freeBoxCount

    # 상품 재입고 필요한 상품 목록 반환한다.
    def NeedReStockList(self, stockManageExel):
        # 해당 월 각 상품 별 하루 판매 평균치 * 10 > 해당 월 재고 수량 시
        self.manageStock = xlwings.Book(stockManageExel)
        self.manageStock.activate()
        needReStock = []

        now = datetime.now()
        nowDate = now.strftime('%Y.%m.')
        sheetName = nowDate[2:8]

        #재입고 필요한 상품 이름, 옵션 리스트 출력
        for i in range(0, ITEMNUMBER):
            sellRocation = 'AJ' + str(i+2)
            stockNameSellRecation = 'A' + str(i+2)
            stockOptionSellRecation = 'B' + str(i+2)
            print(i+2, self.manageStock.sheets[sheetName].range(sellRocation).value)
            if self.manageStock.sheets[sheetName].range(sellRocation).value == '재입고필요':
                stockName = self.manageStock.sheets[sheetName].range(stockNameSellRecation).value
                stockOption = self.manageStock.sheets[sheetName].range(stockOptionSellRecation).value
                #print(stockName, stockOption)
                needReStock.append(stockName)
                if stockOption == None:
                    needReStock.append('')
                else:
                    needReStock.append(stockOption)

        print(needReStock)

        return needReStock

if __name__ == "__main__":
    a = OnlineShopTask()
    #a.NaverPackageDelivery("C:/Users/MINSEUNG KIM/Desktop/매출주문자료/2019-05-27 네이버 추가배송.xlsx", "C:/Users/MINSEUNG KIM/Desktop/매출주문자료/2019-05-27 택배사 추가배송.xls")
    #a.ManageItems("C:/Users/MINSEUNG KIM/Desktop/매출주문자료/2019-05-23 당일배송.xlsx", 'C:/Users/MINSEUNG KIM/Desktop/매출주문자료/재고관리.xlsx')
    #a.ManageFreeItem('C:/Users/MINSEUNG KIM/Desktop/매출주문자료/2019-05-23 당일배송.xlsx', 'C:/Users/MINSEUNG KIM/Desktop/매출주문자료/재고관리.xlsx', nikeShoesList, freeBoxList[0])
    #a.NeedReStockList('C:/Users/MINSEUNG KIM/Desktop/매출주문자료/재고관리.xlsx')